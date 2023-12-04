import csv

from openpyxl import load_workbook

def get_phones_for_calltouch():
    wb2 = load_workbook('Calltouch_2019-05-27_2023-08-20_Журнал_звонков_no_filter.xlsx')
    ws = wb2.active
    set_numbers = set()
    for phone in ws['B']:
        set_numbers.add(phone.value)
    with open('sspot_caltouch_phones.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        id = 1
        writer.writerow(['id','phone'])
        for phone in set_numbers:
            writer.writerow([id, phone])
            id += 1

def get_emails_batiskaf():
    wb2 = load_workbook('Клиенты2108.xlsx')
    ws = wb2.active
    set_emails = set()
    for email in ws['C']:
        set_emails.add(email.value)
    with open('batiskaf_email.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['email'])
        for email in set_emails:
            writer.writerow([email])
get_emails_batiskaf()
get_phones_for_calltouch()